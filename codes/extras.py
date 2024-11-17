import pandas as pd
import os
from datetime import datetime
import calendar
from PySide6.QtWidgets import QApplication, QWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QHeaderView, QLabel, QCalendarWidget, QCheckBox, QMessageBox
from PySide6.QtCore import Qt
from PySide6.QtWidgets import QDateEdit
from PySide6.QtCore import QDate
from PySide6.QtGui import QPixmap  # Import QPixmap for handling images
from num2words import num2words  # We will use this package to convert numbers to words
from openpyxl import load_workbook  # Import openpyxl for handling templates
from PySide6.QtWidgets import QApplication, QWidget, QTableWidget, QTableWidgetItem, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QHeaderView, QLabel, QCalendarWidget, QCheckBox, QMessageBox, QGroupBox
from PySide6.QtCore import Qt, QDate
from PySide6.QtGui import QPixmap, QColor
import pyperclip  # Required for clipboard functionality
import time
import openpyxl
import pygetwindow as gw
import time
import pygetwindow as gw
import ctypes
import screeninfo
import pandas as pd
import pandas as pd
import os

import extras
import my_gui
import main_processing






def load_employee_data(filepath):
    try:
        if not os.path.exists(filepath):
            extras.print_colored(f"File '{filepath}' not found.", "red")
            return None
        
        # Use ExcelFile for explicit file handling (opens and closes the file automatically)
        with pd.ExcelFile(filepath) as xls:
            # Load the employee data sheet
            df_employee = pd.read_excel(xls, sheet_name="employee_data")
            
            if df_employee is None:
                extras.print_colored("No employee data found.", "red")
                return None
            
            # Check if essential columns exist in employee_data sheet
            required_columns_employee = ["Employee_Name", "Employee_Address", "PAN_No", "Account_Number", "Bank_Name", "IFSC_Code", "Gender", 
                                         "MailID", "Phone_No", "Project_Name", "Monthly_Salary", "Food_allowance_per_day_amount", 
                                         "Last_Invoice_No", "Template_No", "Email_Send"]

            for col in required_columns_employee:
                if col not in df_employee.columns:
                    extras.print_colored(f"Missing required column in 'employee_data' sheet: {col}", "red")
                    return None

            # Validate that numeric columns contain valid numbers in employee_data sheet
            for col in ["Monthly_Salary", "Food_allowance_per_day_amount"]:
                if not pd.api.types.is_numeric_dtype(df_employee[col]):
                    extras.print_colored(f"Column '{col}' must contain numeric values in 'employee_data' sheet.", "red")
                    return None

            # Fill any missing values with 0 in the employee data
            df_employee.fillna(0, inplace=True)

            # Load the projects data sheet
            df_projects = pd.read_excel(xls, sheet_name="projects")

            # Check if essential columns exist in projects sheet
            required_columns_projects = ["Project_Name", "Project_Address"]

            for col in required_columns_projects:
                if col not in df_projects.columns:
                    extras.print_colored(f"Missing required column in 'projects' sheet: {col}", "red")
                    return None

            # Load the manager_info sheet
            if 'manager_info' not in xls.sheet_names:
                extras.print_colored("Sheet 'manager_info' does not exist in the file.", "red")
                return None
            
            df_manager_info = pd.read_excel(xls, sheet_name="manager_info")

            # Check for 'Manager_Name' and 'Manager_MailID' columns in the 'manager_info' sheet
            if 'Manager_Name' not in df_manager_info.columns:
                extras.print_colored("Missing 'Manager_Name' column in 'manager_info' sheet.", "red")
                return None
            if 'Manager_MailID' not in df_manager_info.columns:
                extras.print_colored("Missing 'Manager_MailID' column in 'manager_info' sheet.", "red")
                return None

            # Ensure the first row contains values in these columns
            if pd.isna(df_manager_info.at[0, 'Manager_Name']) or pd.isna(df_manager_info.at[0, 'Manager_MailID']):
                extras.print_colored("Missing value in the first row of 'Manager_Name' or 'Manager_MailID' in 'manager_info' sheet.", "red")
                return None

            # After verifying, now strip whitespace from all cells in both sheets
            # Strip whitespace for employee data
            df_employee = df_employee.apply(lambda col: col.str.strip() if col.dtype == "object" else col)

            # Strip whitespace for project data
            df_projects = df_projects.apply(lambda col: col.str.strip() if col.dtype == "object" else col)

            # Strip whitespace for manager_info data
            df_manager_info = df_manager_info.apply(lambda col: col.str.strip() if col.dtype == "object" else col)

            # Save the modified data back to the same file
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_employee.to_excel(writer, sheet_name="employee_data", index=False)
                df_projects.to_excel(writer, sheet_name="projects", index=False)
                df_manager_info.to_excel(writer, sheet_name="manager_info", index=False)

            extras.print_colored(f"Excel file processed successfully and saved back to {filepath}.", "green")
            return df_employee  # Optionally return the modified employee data if needed

    except Exception as e:
        extras.print_colored(f"Data Loading Error: Error loading data: {str(e)}", "red")
        extras.show_command_prompt()

        return None







# def strip_excel_whitespace():
#     try:
#         filepath = r"../invoice_data.xlsx"
        
#         # Attempt to load the Excel file with pandas (preserves all sheets)
#         try:
#             with pd.ExcelFile(filepath) as xls:
#                 # Create a dictionary to hold modified dataframes for each sheet
#                 modified_sheets = {}
                
#                 # Iterate over each sheet in the Excel file
#                 for sheet_name in xls.sheet_names:
#                     try:
#                         # Read the sheet into a DataFrame
#                         df = pd.read_excel(xls, sheet_name=sheet_name)
                        
#                         # Apply strip() to every string cell in the dataframe (for string values)
#                         df = df.apply(lambda col: col.str.strip() if col.dtype == "object" else col)
                        
#                         # Store the modified DataFrame
#                         modified_sheets[sheet_name] = df
                    
#                     except Exception as e:
#                         extras.print_colored(f"Error while processing sheet '{sheet_name}': {str(e)}", "red")
#                         continue  # Skip to the next sheet if an error occurs for this sheet
                
#                 # Attempt to save the modified DataFrames back into the same Excel file
#                 try:
#                     with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
#                         for sheet_name, df in modified_sheets.items():
#                             df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
#                     extras.print_colored(f"Excel file processed successfully and saved back to {filepath}.", "green")
                
#                 except Exception as e:
#                     extras.print_colored(f"Error while saving the modified Excel file: {str(e)}", "red")
        
#         except Exception as e:
#             extras.print_colored(f"Error while loading the Excel file: {str(e)}", "red")
    
#     except Exception as e:
#         extras.print_colored(f"Unexpected error: {str(e)}", "red")







def setup_text_change_listeners(table, row, employee_data, payable_days_input, food_days_input, invoide_date):
    # Connect textChanged signals with correct row reference
    payable_days_input.textChanged.connect(
        lambda: update_salary_for_row(table, row, employee_data, payable_days_input, food_days_input, invoide_date)
    )
    food_days_input.textChanged.connect(
        lambda: update_salary_for_row(table, row, employee_data, payable_days_input, food_days_input, invoide_date)
    )



def get_last_day_of_previous_month():
    # Get the current date
    today = QDate.currentDate()

    # Get the first day of the current month (set day to 1)
    first_day_of_current_month = QDate(today.year(), today.month(), 1)

    # Subtract one day to get the last day of the previous month
    last_day_of_previous_month = first_day_of_current_month.addDays(-1)

    return last_day_of_previous_month





def toggle_select_all(table, button):
    """Toggles the select all/deselect all functionality."""
    is_selecting = button.text() == "Deselect All"
    
    # Loop through all rows and set the checkbox accordingly
    for row in range(table.rowCount()):
        checkbox = table.cellWidget(row, 0)
        if checkbox:
            checkbox.setChecked(is_selecting)

    # Toggle button text
    button.setText("Select All" if is_selecting else "Deselect All")



def generate_invoice_no(current_invoice):
    try:
        parts = current_invoice.split("/")
        
        # Check if the current_invoice format is valid
        if len(parts) != 2 or not parts[0].isdigit() or not parts[1].count('-') == 1:
            extras.print_colored(f"Invalid invoice format: {current_invoice}", "red")

            # print(f"Invalid invoice format: {current_invoice}")
            return "Invalid Invoice"
        
        invoice_no = int(parts[0]) + 1
        fiscal_year = parts[1]
        today = datetime.today()
        month = today.month
        year = today.year
        
        # Determine the current fiscal year based on today's month
        if month > 3:  # After March, the fiscal year is the current year to next year
            fiscal_year_today = f"{year % 100}-{(year + 1) % 100}"  # Example: 2024-25 will be 24-25
        else:  # Before April, the fiscal year is last year to this year
            fiscal_year_today = f"{(year - 1) % 100}-{year % 100}"  # Example: 2023-24 will be 23-24
        
        # Check if the fiscal year in the current invoice matches the current fiscal year
        if fiscal_year != fiscal_year_today:
            # If it's not the current fiscal year, increment both the invoice number and fiscal year
            fiscal_year = fiscal_year_today
            invoice_no = 1  # Start from invoice number 1 for the new fiscal year
        
        # Return the new invoice number and fiscal year in the correct format
        return f"{invoice_no:02}/{fiscal_year}"
    
    except Exception as e:
        extras.print_colored("Invoice Number Error", f"Error generating invoice number: {str(e)}", "red")

        # print("Invoice Number Error", f"Error generating invoice number: {str(e)}")
        return "Invalid Invoice"
    

def format_amount_to_words(amount):
    # Convert the number to words in rupees only (no decimals)
    amount_in_words = num2words(amount, to='currency', currency='INR')
    
    # Capitalize each word and remove commas, hyphens, and the word "Rupees"
    formatted_words = ' '.join(word.capitalize() for word in amount_in_words.replace("Rupees", "").replace(",", "").replace("-", " ").split())
    
    # Remove ", Zero Paise" if it appears in the formatted string
    formatted_words = formatted_words.replace(", Zero Paise", "").strip()

    return f"{formatted_words} Only"



def select_deselect_all(table, select):
    for row in range(table.rowCount()):
        checkbox = table.cellWidget(row, 0)
        checkbox.setChecked(select)






# Function to update the salary for a specific row when payable days or food allowance days change
def update_salary_for_row(table, row, employee_data, payable_days_input, food_days_input, invoice_date):
    try:
        # print(f"--- Updating Salary for Row: {row} ---")

        # Get the values from the inputs
        payable_days = payable_days_input.text().strip()
        food_days = food_days_input.text().strip()
        # print(f"Input Payable Days: {payable_days}, Input Food Days: {food_days}")

        if food_days is None or food_days == "":
            food_days = "0"
        
        # print("this is the food days for updating the cell on gui")
        # print(food_days)
        

        # Check if both values are numeric
        if payable_days.isdigit() and food_days.isdigit():
            payable_days = int(payable_days)
            food_days = int(food_days)
            # print(f"Converted Payable Days: {payable_days}, Converted Food Days: {food_days}")

            # Retrieve the employee's name from the table
            employee_name = table.item(row, 1).text()
            # print(f"Employee Name from Table: {employee_name}")

            # Find the employee index in the employee_data DataFrame
            employee_index = employee_data[employee_data["Employee_Name"] == employee_name].index[0]
            # print(f"Employee Index in DataFrame: {employee_index}")

            # Calculate salary using the updated inputs
            salary = calculate_salary(employee_data, payable_days, food_days, employee_index, invoice_date)
            # print(f"Calculated Salary: {salary}")

            # Update the calculated salary in the table
            table.setItem(row, 5, QTableWidgetItem(f"{salary}"))
            # print(f"Updated Table Salary Cell at Row {row}, Column 5 with: {salary}")

        else:
            extras.print_colored("Error: Payable Days or Food Days input is non-numeric", "red")

            # print("Error: Payable Days or Food Days input is non-numeric")

    except Exception as e:
        extras.print_colored(f"Error updating salary for row {row}: {str(e)}", "red")

        # print(f"Error updating salary for row {row}: {str(e)}")





def get_project_address(project_name_input):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook("../invoice_data.xlsx")
        # print(f"Workbook {workbook} opened successfully.")
        
        # Access the 'projects' sheet
        if 'projects' in workbook.sheetnames:
            sheet = workbook['projects']
            # print(f"Accessing 'projects' sheet.")
        else:
            extras.print_colored(f"The sheet 'projects' was not found in the workbook.", "red")

            # raise ValueError(f"The sheet 'projects' was not found in the workbook.")
        
        # Loop through the rows in the sheet to find the project_name
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from row 2 if row 1 is headers
            project_name = row[0]  # Assuming project_name is in the first column
            project_address = row[1]  # Assuming project_address is in the second column
            
            # Match the project name
            if project_name == project_name_input:
                # print(f"Project '{project_name_input}' found. Returning address.")
                return project_address
        
        # If project_name is not found
        extras.print_colored(f"Project '{project_name_input}' not found in the sheet.", "red")

        # print(f"Project '{project_name_input}' not found in the sheet.")
        return None  # If no match found

    except Exception as e:
        extras.print_colored(f"An error occurred: {e}", "red")

        # print(f"An error occurred: {e}")
        return None
    
    finally:
        # Close the workbook to release the resources
        workbook.close()  # Ensure the workbook is closed after use
        # print("Workbook closed.")

    

def populate_template(employee_data, employee_index, invoice_date):
    try:
        # Fetch employee data
        employee = employee_data.iloc[employee_index]
        template_no = employee["Template_No"]
        template_path = os.path.join("../templates", f"{template_no}.xlsx")

        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file '{template_path}' not found.")


        # print(list(employee_data.columns))
        # Load the workbook using openpyxl
        workbook = load_workbook(template_path)
        sheet = workbook.active  # Assuming the relevant data is in the first sheet

        invoice_data_formatted = invoice_date.toString("yyyy-MM-dd")
        employee_data["Last_Invoice_Date"] = invoice_data_formatted

        # Extract Month and Year from invoice date
        month_name = invoice_date.toString("MMMM")  # "January", "February", etc.
        year = invoice_date.toString("yyyy")  # Get the year
        
        month_year = f"{month_name} - {year}"
        employee_data["Month"] = month_year

        total_salary = employee["Total_Salary"]
        # print("this is the total salaryyyyyyy")
        # print(total_salary)

        # Convert total amount to words (Indian Rupees)
        total_amount_in_words = extras.format_amount_to_words(total_salary)

        # print(total_amount_in_words)

        # Modify the employee address to include line breaks
        employee_address = employee["Employee_Address"].replace("(enter)", "\n")

        employee_name = employee['Employee_Name']
        # print("this is the fucking employee adressssssss")

        # print(employee_address)
        project_address = get_project_address(employee["Project_Name"])
        formatted_project_address = project_address.replace("(enter)", "\n")
        
        # print("this is the fucking project adressssssss")

        # print(formatted_project_address)

        # Prepare the placeholder map
        placeholder_map = {
            "{Enter Employee Name}": employee_name,
            "{Enter Employee Address}": employee_address,
            "{Enter PAN No.}": employee["PAN_No"],
            "{Enter Account Number}": employee["Account_Number"],
            "{Enter Bank Name}": employee["Bank_Name"],
            "{Enter IFSC Code}": employee["IFSC_Code"],
            "{Enter Project Name}": employee["Project_Name"],
            "{Enter Project Address}": formatted_project_address,
            "{Enter Monthly Stipend}": employee["Monthly_Salary"],
            "{Enter Food allowance per day amount}": employee["Food_allowance_per_day_amount"],
            "{Enter Invoice No}": employee["Last_Invoice_No"],
            "{Invoice Date}": invoice_data_formatted,
            "{Month Year}": month_year,  # Added Month and Year
            "{Payable Days}": employee["Payable_Days"],
            "{Payable Days Amount}": employee["Payable_Days_Amount"],
            "{Food Allowance Days Provided}": employee["Food_Days"],
            "{Food Allowance Amount}": employee["Food_Amount"],
            "{Total Amount to be Credited}": total_salary,
            "{Total Amount to be Credited (in Words)}": total_amount_in_words,
        }

        # Replace placeholders in the template
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value in placeholder_map:
                    cell.value = placeholder_map[cell.value]

        # Output path and filename
        output_folder = "../output"
        invoice_folder = os.path.join(output_folder, invoice_data_formatted)
        os.makedirs(invoice_folder, exist_ok=True)

        output_filename = os.path.join(invoice_folder, f"invoice_{employee_name}.xlsx")
        workbook.save(output_filename)
        workbook.close()  # Close the workbook to avoid file locking issues


        extras.print_colored(f"\nTemplate Made succeffully for {employee_index}.{employee_name}", "green")


    except Exception as e:
        extras.print_colored("Template Handling Error", f"Error processing template: {str(e)}", "red")

        # print("Template Handling Error", f"Error processing template: {str(e)}")




def update_email_status(employee_name, status, column_name):
    try:
        # Load the Excel file with all sheets into a dictionary of DataFrames
        filepath = '../invoice_data.xlsx'
        # print(f"Opening Excel file: {filepath}")
        
        # Load the entire workbook into a dictionary of DataFrames (key: sheet name, value: DataFrame)
        all_sheets = pd.read_excel(filepath, sheet_name=None)  # sheet_name=None loads all sheets into a dictionary
        # print("Excel file loaded successfully.")
        
        # Access the 'employee_data' sheet (DataFrame) from the dictionary
        if 'employee_data' not in all_sheets:
            extras.print_colored("'employee_data' sheet not found in the workbook.", "red")

            # raise ValueError("'employee_data' sheet not found in the workbook.")
        
        df_employee_data = all_sheets['employee_data']
        
        # Check if the column exists in the employee_data sheet
        if column_name not in df_employee_data.columns:
            extras.print_colored(f"'{column_name}' column not found in 'employee_data' sheet.", "red")
            # raise ValueError(f"'{column_name}' column not found in 'employee_data' sheet.")
        
        # Locate the row for the employee and update the specified column
        if employee_name in df_employee_data['Employee_Name'].values:
            df_employee_data.loc[df_employee_data['Employee_Name'] == employee_name, column_name] = status
            print(f"{column_name} status for {employee_name} updated to {status}")
        else:
            extras.print_colored(f"Employee '{employee_name}' not found.", "red")

            # print(f"Employee '{employee_name}' not found.")
            return
        
        # Now save all sheets back to the Excel file, with the modified 'employee_data' sheet
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for sheet_name, df in all_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)  # Write each DataFrame to the corresponding sheet
            # print("Excel file saved successfully.")
        
    except Exception as e:
        extras.print_colored(f"Error updating {column_name} for {employee_name}: {str(e)}", "red")

        # print(f"Error updating {column_name} for {employee_name}: {str(e)}")






def show_command_prompt():
    # Set desired Command Prompt window dimensions
    cmd_width = 1200  # Width of the CMD window
    cmd_height = 600  # Height of the CMD window

    # Get all windows with any title
    windows = gw.getWindowsWithTitle("")

    for window in windows:
        if 'cmd' in window.title.lower() or 'python' in window.title.lower():  # Match cmd or python titles
            # Bring the window to the front
            window.activate()
            
            # Get screen dimensions
            screen = screeninfo.get_monitors()[0]  # Assuming primary monitor
            screen_width = screen.width
            screen_height = screen.height

            # Calculate the new position to center the window
            x_pos = (screen_width - cmd_width) // 2
            y_pos = (screen_height - cmd_height) // 2

            # Resize and move the window to the specified size and position
            window.resizeTo(cmd_width, cmd_height)
            window.moveTo(x_pos, y_pos)
            
            break
    else:
        print("Command Prompt window not found.")  # Or use a colored print function if defined





def print_colored(text, color):
    # Set color code based on input
    if color.lower() == "red":
        color_code = "\033[91m"
    elif color.lower() == "green":
        color_code = "\033[92m"
    else:
        color_code = "\033[0m"  # Default color if color input is invalid
    
    # Print the text in the specified color
    print(f"{color_code}{text}\033[0m")









def get_days_in_month(invoice_date):
    # Extract the month and year from the invoice date
    month = invoice_date.month()
    year = invoice_date.year()
    
    # Get the number of days in the month for that year
    days_in_month = QDate(year, month, 1).daysInMonth()  # Using the 1st of the month to get the days
    
    return days_in_month



def calculate_salary(employee_data, payable_days, food_days, employee_index, invoice_date):

    # print("invoice dateeeeeeeeeeeee", invoice_date)
    try:
        # employee = employee_data.loc[employee_index]
        
        month_days = get_days_in_month(invoice_date)
        # print("month days",month_days)

        foodallowanceperdayamount = employee_data.loc[employee_index, "Food_allowance_per_day_amount"]
        # print("foodallowanceperdayamount",foodallowanceperdayamount)

        employeemonthlysalary = employee_data.loc[employee_index, "Monthly_Salary"]
        # print("employeemonthlysalary",employeemonthlysalary)

        foodpart = round( foodallowanceperdayamount * food_days)
        # print("foodpart",foodpart)

        # print("payable_days",payable_days)

        payabledayspart = round(employeemonthlysalary / month_days * payable_days)
        # print("payabledayspart",payabledayspart)

        salary =  foodpart + payabledayspart
        # print("salary",salary)



        # salary = round(salary)
        employee_data.loc[employee_index,"Food_Amount"] = foodpart
        employee_data.loc[employee_index,"Payable_Days_Amount"] = payabledayspart
        employee_data.loc[employee_index,"Total_Salary"] =  salary
        employee_data.loc[employee_index,"Food_Days"] =  food_days
        employee_data.loc[employee_index,"Payable_Days"] =  payable_days


        return salary
    
    except Exception as e:
        extras.print_colored("Calculation Error", f"Error calculating salary: {str(e)}", "red")
        # print("Calculation Error", f"Error calculating salary: {str(e)}")
        return 0.0
    




def get_manager_info():
    try:
        # Load the Excel file
        xl = pd.ExcelFile(r"../invoice_data.xlsx")

        # Check if the 'manager_info' sheet exists
        if 'manager_info' not in xl.sheet_names:
            raise ValueError("Sheet 'manager_info' does not exist in the file.")
        
        # Load the 'manager_info' sheet into a DataFrame
        manager_info_df = xl.parse('manager_info')

        # Check if the 'Manager_Name' column exists
        if 'Manager_Name' not in manager_info_df.columns:
            raise ValueError("Column 'Manager_Name' does not exist in the 'manager_info' sheet.")
        
        # Check if the 'Manager_MailID' column exists
        if 'Manager_MailID' not in manager_info_df.columns:
            raise ValueError("Column 'Manager_MailID' does not exist in the 'manager_info' sheet.")
        
        # Get the values from the second row (index 1)
        manager_name = manager_info_df.at[0, 'Manager_Name']
        manager_email = manager_info_df.at[0, 'Manager_MailID']
        
        # print(manager_name)
        # print(manager_email)

        # Return the values
        return manager_name, manager_email

    except Exception as e:
        print(f"Error: {e}")
        return None, None
